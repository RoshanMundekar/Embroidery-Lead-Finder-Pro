"""
Export service — generates JSON, CSV, and XLSX files from lead data.
"""

import json
from pathlib import Path
from datetime import datetime

import pandas as pd

from app.config import EXPORTS_DIR
from app.utils.logger import logger


def _leads_to_dataframe(leads: list[dict]) -> pd.DataFrame:
    """Convert leads list to a clean Pandas DataFrame."""
    rows = []
    for lead in leads:
        rows.append({
            "Company Name": lead.get("company_name", ""),
            "Website": lead.get("website", ""),
            "Phone": lead.get("phone", ""),
            "Email": lead.get("email", ""),
            "Address": lead.get("address", ""),
            "City": lead.get("city", ""),
            "State": lead.get("state", ""),
            "Rating": lead.get("rating", ""),
            "Reviews": lead.get("reviews", ""),
            "Category": lead.get("category", ""),
            "Facebook": lead.get("social_links", {}).get("facebook", ""),
            "Instagram": lead.get("social_links", {}).get("instagram", ""),
            "LinkedIn": lead.get("social_links", {}).get("linkedin", ""),
        })
    return pd.DataFrame(rows)


def export_to_json(leads: list[dict], search_keyword: str = "", location: str = "") -> str:
    """Export leads to a pretty-printed JSON file. Returns the filepath."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"leads_{timestamp}.json"
    filepath = EXPORTS_DIR / filename

    export_data = {
        "export_date": datetime.now().isoformat(),
        "search_keyword": search_keyword,
        "location": location,
        "total_leads": len(leads),
        "leads": leads,
    }

    with open(filepath, "w", encoding="utf-8") as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)

    logger.info(f"Exported {len(leads)} leads to JSON: {filename}")
    return str(filepath)


def export_to_csv(leads: list[dict], search_keyword: str = "", location: str = "") -> str:
    """Export leads to a CSV file. Returns the filepath."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"leads_{timestamp}.csv"
    filepath = EXPORTS_DIR / filename

    df = _leads_to_dataframe(leads)
    df.to_csv(filepath, index=False, encoding="utf-8-sig")

    logger.info(f"Exported {len(leads)} leads to CSV: {filename}")
    return str(filepath)


def export_to_xlsx(leads: list[dict], search_keyword: str = "", location: str = "") -> str:
    """Export leads to an Excel file with styled headers. Returns the filepath."""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"leads_{timestamp}.xlsx"
    filepath = EXPORTS_DIR / filename

    df = _leads_to_dataframe(leads)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")

        # Style the header row
        worksheet = writer.sheets["Leads"]
        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

        for col_num, col_letter in enumerate(worksheet.iter_cols(min_row=1, max_row=1)):
            for cell in col_letter:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            col_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 4, 50)
            worksheet.column_dimensions[col_letter].width = adjusted_width

    logger.info(f"Exported {len(leads)} leads to XLSX: {filename}")
    return str(filepath)
